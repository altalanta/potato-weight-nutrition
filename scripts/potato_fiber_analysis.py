#!/usr/bin/env python3
"""
Robust Python 3.11 script for processing potato fiber analysis data from Excel files.
Recursively processes .xls/.xlsx files, extracts weight/energy/mood measurements
and nutrition/fiber data, and outputs clean datasets.
"""

import argparse
import logging
import re
import sys
from pathlib import Path

import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook


def setup_logging() -> logging.Logger:
    """Set up logging configuration."""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler('potato_processing.log')
        ]
    )
    return logging.getLogger(__name__)


def normalize_text(text: str) -> str:
    """Normalize text: strip, lower, collapse whitespace."""
    if pd.isna(text) or not isinstance(text, str):
        return ""
    return re.sub(r'\s+', ' ', str(text).strip().lower())


def coerce_numeric(series: pd.Series) -> pd.Series:
    """
    Parse numbers from messy strings like "12 g", "~8.5", "3,221", "5g/day".
    Keep NaN for blank/none.
    """
    def parse_numeric(val):
        if pd.isna(val):
            return pd.NA
        
        if isinstance(val, (int, float)):
            return float(val) if not pd.isna(val) else pd.NA
        
        # Convert to string and clean
        val_str = str(val).strip()
        if not val_str or val_str.lower() in ['', 'nan', 'none', 'null']:
            return pd.NA
            
        # Remove common prefixes and suffixes, keep numbers and decimal points
        # Handle cases like "~8.5", "12 g", "3,221", "5g/day"
        cleaned = re.sub(r'[^\d.,\-+]', '', val_str)
        cleaned = cleaned.replace(',', '')  # Remove thousands separators
        
        if not cleaned:
            return pd.NA
            
        try:
            return float(cleaned)
        except ValueError:
            return pd.NA
    
    return series.apply(parse_numeric)


def trailing_int(s: str) -> int | None:
    """Extract trailing integer from string (for headers like weight_15)."""
    if pd.isna(s) or not isinstance(s, str):
        return None
    
    match = re.search(r'(\d+)$', s.strip())
    return int(match.group(1)) if match else None


def safe_read_excel(file_path: Path, sheet_name: str, header: int | None = 0) -> pd.DataFrame | None:
    """Safely read Excel sheet with error handling."""
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header)
        return df if not df.empty else None
    except Exception as e:
        logging.warning(f"Failed to read {file_path}:{sheet_name} with header={header}: {e}")
        return None


def sniff_header_row(df_no_header: pd.DataFrame) -> int:
    """
    Find plausible header row by choosing first row with ≥2 distinct non-null values
    and at least one string.
    """
    for i, row in df_no_header.iterrows():
        non_null_values = row.dropna()
        if len(non_null_values) >= 2:
            unique_values = non_null_values.unique()
            if len(unique_values) >= 2 and any(isinstance(v, str) for v in unique_values):
                return i
    return 0  # Fallback to first row


def detect_label_style_sheet(df: pd.DataFrame) -> bool:
    """
    Detect if sheet has label-style layout (first column contains row labels).
    Look for ≥6 rows where first column matches subject id, weight, energy, mood patterns.
    """
    if df.empty or len(df) < 6:
        return False
    
    first_col = df.iloc[:, 0].astype(str).str.lower().str.strip()
    
    patterns = [
        r'subject\s*id',
        r'starting\s*(weight|energy|mood)',
        r'(weight|energy|mood)\s*end\s*week',
        r'energy\s*end\s*week',
        r'mood\s*end\s*week',
        r'end\s*60\s*days?'
    ]
    
    matches = 0
    for pattern in patterns:
        if any(re.search(pattern, str(cell)) for cell in first_col):
            matches += 1
    
    return matches >= 3  # At least 3 pattern matches


def tidy_label_style_sheet(df: pd.DataFrame, sheet_name: str, source_path: Path) -> pd.DataFrame:
    """
    Process label-style sheet into tidy format.
    Returns DataFrame with columns: subject_id, metric, phase, week, value, tp_index, timepoint, note, sheet_name, source_path
    """
    if df.empty:
        return pd.DataFrame()
    
    # Rename first column to 'label'
    df = df.copy()
    df.columns = ['label'] + [f'timepoint_{i}' if pd.isna(col) or str(col).startswith('Unnamed') 
                              else str(col) for i, col in enumerate(df.columns[1:], 1)]
    
    # Melt to long format
    id_vars = ['label']
    value_vars = [col for col in df.columns if col != 'label']
    melted = pd.melt(df, id_vars=id_vars, value_vars=value_vars, 
                     var_name='timepoint', value_name='value')
    
    # Extract subject_id
    subject_id_row = melted[melted['label'].str.contains(r'subject\s*id', case=False, na=False)]
    subject_id = None
    if not subject_id_row.empty:
        subject_id = subject_id_row['value'].dropna().iloc[0] if not subject_id_row['value'].dropna().empty else None
    
    # Parse labels
    results = []
    for _, row in melted.iterrows():
        label = normalize_text(str(row['label']))
        value = row['value']
        timepoint = row['timepoint']
        
        # Skip subject_id rows
        if re.search(r'subject\s*id', label):
            continue
            
        # Parse label patterns
        metric = None
        phase = None
        week = None
        note = None
        
        # Starting patterns
        start_match = re.search(r'starting\s*(weight|energy|mood)', label)
        if start_match:
            metric = start_match.group(1)
            phase = 'start'
            week = 0
        
        # End week patterns
        end_week_match = re.search(r'(weight|energy|mood)\s*end\s*week\s*(\d+)', label)
        if end_week_match:
            metric = end_week_match.group(1)
            phase = 'end'
            week = int(end_week_match.group(2))
        
        # End 60 days pattern
        end_60_match = re.search(r'(weight|energy|mood)\s*end\s*60\s*days?', label)
        if end_60_match:
            metric = end_60_match.group(1)
            phase = 'end'
            week = 60
            note = '60d'
        
        # Skip if no pattern matched
        if metric is None:
            continue
        
        # Convert value to numeric
        numeric_value = coerce_numeric(pd.Series([value])).iloc[0]
        
        # Treat zeros as missing for phase=end rows
        if phase == 'end' and numeric_value == 0:
            numeric_value = pd.NA
        
        # Extract tp_index from timepoint
        tp_index = trailing_int(timepoint)
        
        results.append({
            'subject_id': subject_id,
            'metric': metric,
            'phase': phase,
            'week': week,
            'value': numeric_value,
            'tp_index': tp_index,
            'timepoint': timepoint,
            'note': note,
            'sheet_name': sheet_name,
            'source_path': str(source_path)
        })
    
    return pd.DataFrame(results)


def find_nutrition_tables(df: pd.DataFrame) -> bool:
    """
    Check if DataFrame contains nutrition/fiber information.
    Returns True if any column header contains fiber or nutrition context hints.
    """
    if df.empty:
        return False
    
    # Get column headers as normalized strings
    headers = [normalize_text(str(col)) for col in df.columns]
    header_text = ' '.join(headers)
    
    # Fiber synonyms
    fiber_terms = [
        'fiber', 'fibre', 'dietary fiber', 'dietary fibre', 'total fiber',
        'total dietary fiber', 'df', 'tdf', 'insoluble', 'soluble'
    ]
    
    # Check for fiber terms
    has_fiber = any(term in header_text for term in fiber_terms)
    
    # Nutrition context hints (boost confidence)
    nutrition_hints = [
        'food', 'item', 'meal', 'date', 'day', 'serving', 'brand',
        'calories', 'carb', 'protein', 'fat', 'net carb', 'sugar'
    ]
    
    has_nutrition_context = any(hint in header_text for hint in nutrition_hints)
    
    return has_fiber or (has_nutrition_context and len([h for h in headers if 'g' in h or 'mg' in h]) > 0)


def standardize_nutrition_frame(df: pd.DataFrame, sheet_name: str, source_path: Path, 
                               workbook_subject_id: str | None = None) -> pd.DataFrame:
    """
    Standardize nutrition DataFrame to consistent format.
    Returns DataFrame with columns: subject_id, date, food, fiber_g, calories_kcal, carb_g, protein_g, fat_g, sugar_g, insoluble_fiber_g, soluble_fiber_g, sheet_name, source_path
    """
    if df.empty:
        return pd.DataFrame()
    
    df = df.copy()
    
    # Normalize column names  
    normalized_cols = []
    for col in df.columns:
        norm_col = normalize_text(str(col))
        # Replace various punctuation with underscores but preserve the normalized text
        norm_col = norm_col.replace(' ', '_').replace('-', '_').replace('.', '_').replace('(', '').replace(')', '')
        normalized_cols.append(norm_col)
    df.columns = normalized_cols
    
    # Remove duplicate/empty column names
    new_cols = []
    for col in df.columns:
        if col in new_cols or not col:
            new_cols.append(f"unnamed_{len(new_cols)}")
        else:
            new_cols.append(col)
    df.columns = new_cols
    
    results = []
    
    for _, row in df.iterrows():
        record = {}
        
        # Date parsing
        date_val = None
        date_cols = [col for col in df.columns if any(term in col for term in ['date', 'day'])]
        if date_cols:
            date_raw = row[date_cols[0]]
            try:
                if pd.notna(date_raw):
                    if isinstance(date_raw, (int, float)):
                        # Handle Excel serial dates
                        date_val = pd.to_datetime('1899-12-30') + pd.Timedelta(days=date_raw)
                    else:
                        date_val = pd.to_datetime(date_raw, errors='coerce')
            except:
                date_val = pd.NaT
        
        # Food/item name
        food_val = None
        food_cols = [col for col in df.columns if any(term in col for term in ['food', 'item', 'meal', 'description', 'entry', 'name'])]
        if food_cols:
            food_val = str(row[food_cols[0]]) if pd.notna(row[food_cols[0]]) else None
        
        # Fiber calculation - prefer specific over generic
        fiber_g = pd.NA
        insoluble_g = pd.NA
        soluble_g = pd.NA
        
        # Look for specific fiber components first
        insol_cols = [col for col in df.columns if 'insoluble' in col]
        sol_cols = [col for col in df.columns if 'soluble' in col]
        
        if insol_cols:
            insoluble_g = coerce_numeric(pd.Series([row[insol_cols[0]]])).iloc[0]
        if sol_cols:
            soluble_g = coerce_numeric(pd.Series([row[sol_cols[0]]])).iloc[0]
        
        # If we have both components, sum them
        if pd.notna(insoluble_g) and pd.notna(soluble_g):
            fiber_g = insoluble_g + soluble_g
        else:
            # Look for generic fiber columns
            fiber_cols = [col for col in df.columns if any(term in col for term in [
                'dietary_fiber', 'total_dietary_fiber', 'total_fiber', 'fiber', 'fibre', 'tdf', 'df'
            ]) and not any(comp in col for comp in ['insoluble', 'soluble'])]
            
            if fiber_cols:
                fiber_raw = row[fiber_cols[0]]
                fiber_g = coerce_numeric(pd.Series([fiber_raw])).iloc[0]
                
                # Convert mg to g if units detected
                if isinstance(fiber_raw, str) and 'mg' in str(fiber_raw).lower():
                    if pd.notna(fiber_g):
                        fiber_g = fiber_g / 1000
        
        # Other nutrients
        nutrients = {}
        nutrient_map = {
            'calories_kcal': ['calories', 'kcal', 'cal'],
            'carb_g': ['carb', 'carbohydrate'],
            'protein_g': ['protein'],
            'fat_g': ['fat'],
            'sugar_g': ['sugar']
        }
        
        for nutrient_key, search_terms in nutrient_map.items():
            nutrient_cols = [col for col in df.columns if any(term in col for term in search_terms)]
            if nutrient_cols:
                nutrients[nutrient_key] = coerce_numeric(pd.Series([row[nutrient_cols[0]]])).iloc[0]
            else:
                nutrients[nutrient_key] = pd.NA
        
        # Subject ID
        subject_id = workbook_subject_id
        subj_cols = [col for col in df.columns if 'subject' in col and 'id' in col]
        if subj_cols and pd.notna(row[subj_cols[0]]):
            subject_id = str(row[subj_cols[0]])
        
        record = {
            'subject_id': subject_id,
            'date': date_val,
            'food': food_val,
            'fiber_g': fiber_g,
            'calories_kcal': nutrients['calories_kcal'],
            'carb_g': nutrients['carb_g'],
            'protein_g': nutrients['protein_g'],
            'fat_g': nutrients['fat_g'],
            'sugar_g': nutrients['sugar_g'],
            'insoluble_fiber_g': insoluble_g,
            'soluble_fiber_g': soluble_g,
            'sheet_name': sheet_name,
            'source_path': str(source_path)
        }
        
        results.append(record)
    
    return pd.DataFrame(results)


def rollup_fiber_daily(nutrition_df: pd.DataFrame) -> pd.DataFrame:
    """
    Create daily rollup of fiber and nutrients by subject_id and date.
    Returns DataFrame with columns: subject_id, date, fiber_g_total, calories_kcal_total, carb_g_total, protein_g_total, fat_g_total, n_entries
    """
    if nutrition_df.empty:
        return pd.DataFrame()
    
    # Filter out rows with missing dates
    df_with_dates = nutrition_df[nutrition_df['date'].notna()].copy()
    
    if df_with_dates.empty:
        return pd.DataFrame()
    
    # Group by subject_id and date
    grouped = df_with_dates.groupby(['subject_id', 'date'])
    
    # Aggregate
    agg_dict = {
        'fiber_g': 'sum',
        'calories_kcal': 'sum',
        'carb_g': 'sum',
        'protein_g': 'sum',
        'fat_g': 'sum'
    }
    
    daily_totals = grouped.agg(agg_dict).reset_index()
    daily_totals['n_entries'] = grouped.size().values
    
    # Rename columns to match expected output
    daily_totals.columns = ['subject_id', 'date', 'fiber_g_total', 'calories_kcal_total', 
                           'carb_g_total', 'protein_g_total', 'fat_g_total', 'n_entries']
    
    return daily_totals


def find_excels(base_path: Path) -> list[Path]:
    """Find all .xls/.xlsx files in base directory and up to 3 immediate subfolders."""
    excel_files = []
    
    # Files in base directory
    for pattern in ['*.xls', '*.xlsx']:
        excel_files.extend(base_path.glob(pattern))
    
    # Files in immediate subdirectories (up to 3)
    subdirs = [d for d in base_path.iterdir() if d.is_dir()][:3]
    for subdir in subdirs:
        for pattern in ['*.xls', '*.xlsx']:
            excel_files.extend(subdir.glob(pattern))
    
    return sorted(excel_files)


def process_workbook(file_path: Path, logger: logging.Logger) -> tuple[list[pd.DataFrame], list[pd.DataFrame]]:
    """
    Process a single Excel workbook.
    Returns tuple of (measures_data, nutrition_data) lists.
    """
    measures_data = []
    nutrition_data = []
    workbook_subject_id = None
    
    try:
        # Get sheet names
        try:
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        except Exception as e:
            logger.warning(f"Could not read sheet names from {file_path}: {e}")
            # Fallback: try reading with pandas
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            excel_file.close()
        
        logger.info(f"Processing {file_path} with {len(sheet_names)} sheets")
        
        for sheet_name in sheet_names:
            try:
                # Try header=0 first
                df = safe_read_excel(file_path, sheet_name, header=0)
                header_strategy = "header=0"
                
                # Check if we need to infer header
                if df is not None and (
                    df.empty or 
                    df.iloc[:, 0].isna().sum() > len(df) * 0.8 or  # First column mostly NaN
                    len([col for col in df.columns if str(col).startswith('Unnamed')]) > len(df.columns) * 0.5
                ):
                    # Try reading without header and infer
                    df_no_header = safe_read_excel(file_path, sheet_name, header=None)
                    if df_no_header is not None:
                        header_row = sniff_header_row(df_no_header)
                        df = safe_read_excel(file_path, sheet_name, header=header_row)
                        header_strategy = f"inferred header at row {header_row}"
                
                if df is None or df.empty:
                    logger.warning(f"Empty or unreadable sheet: {file_path}:{sheet_name}")
                    continue
                
                logger.debug(f"Sheet {sheet_name}: {header_strategy}, shape {df.shape}")
                
                # Check for label-style sheet (weight/energy/mood)
                if detect_label_style_sheet(df):
                    tidy_data = tidy_label_style_sheet(df, sheet_name, file_path)
                    if not tidy_data.empty:
                        measures_data.append(tidy_data)
                        # Extract subject_id for workbook context
                        if workbook_subject_id is None and 'subject_id' in tidy_data.columns:
                            subj_vals = tidy_data['subject_id'].dropna()
                            if not subj_vals.empty:
                                workbook_subject_id = str(subj_vals.iloc[0])
                        logger.info(f"Extracted {len(tidy_data)} measures from {sheet_name}")
                
                # Check for nutrition tables
                if find_nutrition_tables(df):
                    nutrition_frame = standardize_nutrition_frame(df, sheet_name, file_path, workbook_subject_id)
                    if not nutrition_frame.empty:
                        nutrition_data.append(nutrition_frame)
                        logger.info(f"Extracted {len(nutrition_frame)} nutrition rows from {sheet_name}")
                
            except Exception as e:
                logger.error(f"Error processing sheet {file_path}:{sheet_name}: {e}")
                continue
    
    except Exception as e:
        logger.error(f"Error processing workbook {file_path}: {e}")
    
    return measures_data, nutrition_data


def save_outputs(measures_all: pd.DataFrame, nutrition_all: pd.DataFrame, 
                fiber_daily_all: pd.DataFrame, output_dir: Path, logger: logging.Logger):
    """Save all outputs to parquet and CSV files."""
    
    # Ensure output columns are in correct order with all required columns present
    measures_columns = ['subject_id', 'metric', 'phase', 'week', 'value', 'tp_index', 'timepoint', 'note', 'sheet_name', 'source_path']
    nutrition_columns = ['subject_id', 'date', 'food', 'fiber_g', 'calories_kcal', 'carb_g', 'protein_g', 'fat_g', 'sugar_g', 'insoluble_fiber_g', 'soluble_fiber_g', 'sheet_name', 'source_path']
    daily_columns = ['subject_id', 'date', 'fiber_g_total', 'calories_kcal_total', 'carb_g_total', 'protein_g_total', 'fat_g_total', 'n_entries']
    
    # Measures data
    if not measures_all.empty:
        # Ensure all columns exist
        for col in measures_columns:
            if col not in measures_all.columns:
                measures_all[col] = pd.NA
        measures_all = measures_all[measures_columns]
        
        # Save
        measures_parquet = output_dir / "Potato_tidy.parquet"
        measures_csv = output_dir / "Potato_tidy.csv"
        
        table = pa.Table.from_pandas(measures_all)
        pq.write_table(table, measures_parquet)
        measures_all.to_csv(measures_csv, index=False)
        
        logger.info(f"Saved {len(measures_all)} measures to {measures_parquet} and {measures_csv}")
    else:
        logger.warning("No measures data to save")
    
    # Nutrition data
    if not nutrition_all.empty:
        # Ensure all columns exist
        for col in nutrition_columns:
            if col not in nutrition_all.columns:
                nutrition_all[col] = pd.NA
        nutrition_all = nutrition_all[nutrition_columns]
        
        # Save
        nutrition_parquet = output_dir / "Potato_nutrition_rows.parquet"
        nutrition_csv = output_dir / "Potato_nutrition_rows.csv"
        
        table = pa.Table.from_pandas(nutrition_all)
        pq.write_table(table, nutrition_parquet)
        nutrition_all.to_csv(nutrition_csv, index=False)
        
        logger.info(f"Saved {len(nutrition_all)} nutrition rows to {nutrition_parquet} and {nutrition_csv}")
    else:
        logger.warning("No nutrition data to save")
    
    # Daily fiber data
    if not fiber_daily_all.empty:
        # Ensure all columns exist
        for col in daily_columns:
            if col not in fiber_daily_all.columns:
                fiber_daily_all[col] = pd.NA
        fiber_daily_all = fiber_daily_all[daily_columns]
        
        # Save
        daily_parquet = output_dir / "Potato_fiber_daily.parquet"
        daily_csv = output_dir / "Potato_fiber_daily.csv"
        
        table = pa.Table.from_pandas(fiber_daily_all)
        pq.write_table(table, daily_parquet)
        fiber_daily_all.to_csv(daily_csv, index=False)
        
        logger.info(f"Saved {len(fiber_daily_all)} daily fiber records to {daily_parquet} and {daily_csv}")
    else:
        logger.warning("No daily fiber data to save")


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(description='Process potato fiber analysis Excel files')
    parser.add_argument('--base', default='~/Downloads/Potato Raw Dato', 
                       help='Base directory to process (default: ~/Downloads/Potato Raw Dato)')
    parser.add_argument('--test', action='store_true', help='Run self-tests only')
    
    args = parser.parse_args()
    
    if args.test:
        run_self_tests()
        return
    
    # Setup
    logger = setup_logging()
    base_path = Path(args.base).expanduser()
    
    if not base_path.exists():
        logger.error(f"Base path does not exist: {base_path}")
        sys.exit(1)
    
    logger.info(f"Starting processing of {base_path}")
    
    # Create output directory
    output_dir = base_path / "_clean"
    output_dir.mkdir(exist_ok=True)
    
    # Find Excel files
    excel_files = find_excels(base_path)
    logger.info(f"Found {len(excel_files)} Excel files to process")
    
    if not excel_files:
        logger.warning("No Excel files found")
        return
    
    # Process all files
    all_measures = []
    all_nutrition = []
    
    for file_path in excel_files:
        try:
            measures_data, nutrition_data = process_workbook(file_path, logger)
            all_measures.extend(measures_data)
            all_nutrition.extend(nutrition_data)
        except Exception as e:
            logger.error(f"Failed to process {file_path}: {e}")
            continue
    
    # Combine all data
    measures_all = pd.concat(all_measures, ignore_index=True) if all_measures else pd.DataFrame()
    nutrition_all = pd.concat(all_nutrition, ignore_index=True) if all_nutrition else pd.DataFrame()
    
    # Create daily fiber rollup
    fiber_daily_all = rollup_fiber_daily(nutrition_all)
    
    # Save outputs
    save_outputs(measures_all, nutrition_all, fiber_daily_all, output_dir, logger)
    
    # Summary
    logger.info("Processing complete:")
    logger.info(f"  - {len(measures_all)} total measures")
    logger.info(f"  - {len(nutrition_all)} total nutrition rows")  
    logger.info(f"  - {len(fiber_daily_all)} daily fiber records")
    logger.info(f"  - Output saved to {output_dir}")


def run_self_tests():
    """Run inline self-tests."""
    print("Running self-tests...")
    
    # Test 1: coerce_numeric function
    print("Testing coerce_numeric...")
    test_series = pd.Series(["12 g", "~8.5", "3,221", "5g/day", "", "nan", 42, 3.14, "850 mg"])
    numeric_result = coerce_numeric(test_series)
    expected = [12.0, 8.5, 3221.0, 5.0, pd.NA, pd.NA, 42.0, 3.14, 850.0]
    
    for i, (result, expect) in enumerate(zip(numeric_result, expected, strict=False)):
        if pd.isna(expect) and pd.isna(result):
            continue
        elif pd.notna(expect) and pd.notna(result) and abs(result - expect) < 0.001:
            continue
        else:
            print(f"  FAIL: Index {i}, got {result}, expected {expect}")
            return False
    print("  PASS: coerce_numeric")
    
    # Test 2: trailing_int function
    print("Testing trailing_int...")
    test_cases = [("weight_15", 15), ("timepoint_3", 3), ("energy", None), ("mood_end_60", 60)]
    for text, expected in test_cases:
        result = trailing_int(text)
        if result != expected:
            print(f"  FAIL: trailing_int('{text}') = {result}, expected {expected}")
            return False
    print("  PASS: trailing_int")
    
    # Test 3: Create toy nutrition frame and test standardization
    print("Testing nutrition frame processing...")
    
    # Create toy data with mixed fiber columns and date formats
    toy_data = {
        'Date': ['2023-01-01', '2023-01-02', 44928, 44929],  # Mix of string and Excel serial dates
        'Food Item': ['Apple', 'Bread', 'Potato', 'Rice'],
        'Fibre (g)': ['2.4 g', '3.0', '2.2', '0.4'],
        'Insoluble Fiber': [1.5, 2.0, 1.8, 0.2],
        'Soluble Fiber': [0.9, 1.0, 0.4, 0.2],
        'Calories': [95, 79, 77, 130],
        'Protein (g)': [0.5, 2.7, 2.0, 2.7],
        'Subject ID': ['SUBJ_001', 'SUBJ_001', 'SUBJ_002', 'SUBJ_002']
    }
    
    toy_df = pd.DataFrame(toy_data)
    
    # Test standardization
    standardized = standardize_nutrition_frame(toy_df, 'test_sheet', Path('/test/path.xlsx'))
    
    if standardized.empty:
        print("  FAIL: Standardized frame is empty")
        return False
    
    
    # Check that fiber_g was calculated from components where available
    if len(standardized) != 4:
        print(f"  FAIL: Expected 4 rows, got {len(standardized)}")
        return False
    
    # For first row, fiber_g should be calculated from components (insoluble + soluble)
    first_fiber = standardized.iloc[0]['fiber_g']
    first_insoluble = standardized.iloc[0]['insoluble_fiber_g']
    first_soluble = standardized.iloc[0]['soluble_fiber_g']
    
    # Check that components are correctly parsed and summed
    if pd.notna(first_insoluble) and pd.notna(first_soluble):
        expected_fiber = first_insoluble + first_soluble
        if abs(first_fiber - expected_fiber) > 0.001:
            print(f"  FAIL: First row fiber_g = {first_fiber}, but insoluble + soluble = {expected_fiber}")
            return False
    else:
        print(f"  FAIL: Expected both insoluble and soluble to be parsed, got {first_insoluble}, {first_soluble}")
        return False
    
    # Check date parsing worked
    if pd.isna(standardized.iloc[0]['date']):
        print("  FAIL: Date parsing failed for first row")
        return False
    
    print("  PASS: nutrition frame standardization")
    
    # Test 4: Daily rollup
    print("Testing daily fiber rollup...")
    
    daily_rollup = rollup_fiber_daily(standardized)
    
    if daily_rollup.empty:
        print("  FAIL: Daily rollup is empty")
        return False
    
    # Should have entries for each subject-date combination
    expected_entries = len(standardized[standardized['date'].notna()].groupby(['subject_id', 'date']))
    if len(daily_rollup) != expected_entries:
        print(f"  FAIL: Expected {expected_entries} daily entries, got {len(daily_rollup)}")
        return False
    
    # Check that fiber totals are reasonable
    if daily_rollup['fiber_g_total'].isna().all():
        print("  FAIL: All fiber totals are NaN")
        return False
    
    print("  PASS: daily fiber rollup")
    
    # Test 5: Label-style detection
    print("Testing label-style sheet detection...")
    
    # Create a toy label-style sheet (need at least 6 rows for detection)
    label_data = {
        'Measure': ['Subject ID', 'Starting Weight', 'Weight End Week 4', 'Energy End Week 8', 'Mood End 60 Days', 'Energy End Week 12'],
        'TP1': ['SUBJ_001', 70.5, 68.2, 7.5, 6.8, 7.2],
        'TP2': ['SUBJ_002', 65.0, 63.1, 8.0, 7.2, 7.8],
        'TP3': [pd.NA, 72.3, 70.0, 6.9, 6.5, 7.0]
    }
    
    label_df = pd.DataFrame(label_data)
    
    is_label_style = detect_label_style_sheet(label_df)
    if not is_label_style:
        print("  FAIL: Failed to detect label-style sheet")
        return False
    
    # Test tidying
    tidy_result = tidy_label_style_sheet(label_df, 'test_sheet', Path('/test/path.xlsx'))
    
    if tidy_result.empty:
        print("  FAIL: Tidy result is empty")
        return False
    
    # Should have extracted weight, energy, and mood measurements
    metrics = tidy_result['metric'].unique()
    expected_metrics = {'weight', 'energy', 'mood'}
    if set(metrics) != expected_metrics:
        print(f"  FAIL: Expected metrics {expected_metrics}, got {set(metrics)}")
        return False
    
    print("  PASS: label-style sheet processing")
    
    print("All self-tests passed! ✓")
    return True


if __name__ == "__main__":
    main()